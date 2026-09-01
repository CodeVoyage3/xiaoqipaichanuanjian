"""Read-only, repeatable aggregate profile and cold-start simulation for V1-F01.

Usage:
  python tools/v1_f01_analyze.py --source <xlsx> --output obj/V1F01/summary.json
"""
import argparse
import collections
import datetime as dt
import hashlib
import json
import math
from pathlib import Path

import openpyxl

BUSINESS_DATE = dt.date(2026, 9, 1)
REQUIRED = ("商品大类", "商品中类", "商品小类", "商品编码", "商品条码", "商品名称", "生产日期", "有效日期", "保质期", "保质期单位", "是否该做临期折扣", "该批次累计到货数量", "该商品门店库存总数")
GENERAL_LONG_LIFE_CATEGORIES = {"日用", "美妆", "家居", "香氛香水", "文具", "潮流玩具"}
V1_EXCLUDED_CATEGORIES = {"应季搭配", "赠品小样"}
SCHEMES = {
    "ratio_3pct": (0.03, None, None),
    "ratio_5pct": (0.05, None, None),
    "ratio_10pct": (0.10, None, None),
    "ratio_5pct_clamp_3_30": (0.05, 3, 30),
    "ratio_5pct_clamp_3_60": (0.05, 3, 60),
    "ratio_5pct_clamp_7_60": (0.05, 7, 60),
    "ratio_3pct_clamp_3_30": (0.03, 3, 30),
}


def clean(value):
    return None if value is None else str(value).strip() or None


def date_value(value):
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return None


def integer(value):
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = clean(value)
    try:
        return int(text) if text and str(int(text)) == text else None
    except ValueError:
        return None


def stage_from_windows(expiry, first, second, third):
    remaining = (expiry - BUSINESS_DATE).days
    if remaining <= 0:
        return "expired"
    if remaining > first:
        return "none"
    if remaining > second:
        return "discount_50"
    if remaining > third:
        return "discount_20"
    return "withdraw"


def window(days, ratio, lower, upper):
    value = math.ceil(days * ratio)
    if lower is not None:
        value = max(value, lower)
    if upper is not None:
        value = min(value, upper)
    return value


def approved_policy(record):
    """Return (policy_code, stage) only where the company rule is fully applicable."""
    category, total = record["category"], record["declared_days"]
    if category == "食品":
        windows = (30, 14, 7) if total <= 270 else (90, 60, 14)
        return "food_expiry_v1", stage_from_windows(record["expiry"], *windows)
    if category == "宠物":
        return "pet_expiry_v1", stage_from_windows(record["expiry"], 90, 60, 14)
    if category in GENERAL_LONG_LIFE_CATEGORIES:
        if total > 180:
            return "general_long_expiry_v1", stage_from_windows(record["expiry"], 180, 90, 14)
        return None, "uncovered_total_shelf_life_le_6_months"
    if category in V1_EXCLUDED_CATEGORIES:
        return None, "excluded_from_expiry_management_in_v1"
    return None, "no_confirmed_policy_for_category"


def percentile_nearest_rank(values, percentile):
    if not values:
        return None
    return sorted(values)[math.ceil(len(values) * percentile) - 1]


def counter_dict(counter):
    return dict(sorted(counter.items(), key=lambda x: str(x[0])))


def main():
    args = argparse.ArgumentParser()
    args.add_argument("--source", required=True)
    args.add_argument("--output", required=True)
    ns = args.parse_args()
    source, output = Path(ns.source), Path(ns.output)
    before = hashlib.sha256(source.read_bytes()).hexdigest().upper()
    size = source.stat().st_size
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    formula_wb = openpyxl.load_workbook(source, read_only=True, data_only=False)
    sheets = []
    for ws in wb.worksheets:
        headers = [clean(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        formulas = sum(1 for row in formula_wb[ws.title].iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("="))
        sheets.append({"name": ws.title, "rows": ws.max_row, "columns": ws.max_column, "headers": headers, "formula_count": formulas})
    if len(wb.worksheets) != 1:
        raise SystemExit("Expected exactly one worksheet for the V1-F01 source.")
    ws = wb.active
    headers = sheets[0]["headers"]
    missing = [x for x in REQUIRED if x not in headers]
    if missing:
        raise SystemExit("Missing required headers: " + ", ".join(missing))
    col = {header: headers.index(header) for header in REQUIRED}
    rows = []
    for excel_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = {name: values[index] if index < len(values) else None for name, index in col.items()}
        record["row"] = excel_row
        rows.append(record)

    field_quality = {name: collections.Counter() for name in REQUIRED}
    category_rows, product_rows, barcode_rows, batch_rows = collections.defaultdict(list), collections.defaultdict(list), collections.defaultdict(list), collections.defaultdict(list)
    usable = []
    invalid_reasons = collections.Counter()
    for r in rows:
        category, middle, small = clean(r["商品大类"]), clean(r["商品中类"]), clean(r["商品小类"])
        code, barcode, name = clean(r["商品编码"]), clean(r["商品条码"]), clean(r["商品名称"])
        production, expiry = date_value(r["生产日期"]), date_value(r["有效日期"])
        life, unit, stock, arrival = integer(r["保质期"]), clean(r["保质期单位"]), integer(r["该商品门店库存总数"]), integer(r["该批次累计到货数量"])
        for field, value in (("商品大类", category), ("商品中类", middle), ("商品小类", small), ("商品编码", code), ("商品条码", barcode), ("商品名称", name), ("生产日期", production), ("有效日期", expiry), ("保质期", life), ("保质期单位", unit), ("是否该做临期折扣", clean(r["是否该做临期折扣"])), ("该批次累计到货数量", arrival), ("该商品门店库存总数", stock)):
            field_quality[field]["missing" if value is None else "present"] += 1
        if r["生产日期"] is not None and production is None: field_quality["生产日期"]["invalid"] += 1
        if r["有效日期"] is not None and expiry is None: field_quality["有效日期"]["invalid"] += 1
        if r["保质期"] is not None and life is None: field_quality["保质期"]["invalid"] += 1
        if life is not None and life <= 0: field_quality["保质期"]["non_positive"] += 1
        if unit is not None and unit not in {"D", "M", "Y"}: field_quality["保质期单位"]["invalid"] += 1
        if stock is not None and stock < 0: field_quality["该商品门店库存总数"]["negative"] += 1
        if arrival is not None and arrival < 0: field_quality["该批次累计到货数量"]["negative"] += 1
        category_rows[category or "<空白>"].append(r)
        if code: product_rows[code].append(r)
        if barcode: barcode_rows[barcode].append(r)
        key = (code, production, expiry)
        if code and expiry: batch_rows[key].append(r)
        reasons = []
        if not code: reasons.append("missing_product_code")
        if not expiry: reasons.append("missing_or_invalid_expiry_date")
        if life is None or life <= 0: reasons.append("missing_invalid_or_nonpositive_shelf_life")
        if unit not in {"D", "M", "Y"}: reasons.append("missing_or_invalid_shelf_life_unit")
        if stock is None or stock < 0: reasons.append("missing_invalid_or_negative_stock")
        if production is None: reasons.append("missing_or_invalid_production_date_for_actual_days")
        if production and expiry and expiry < production: reasons.append("expiry_before_production")
        if reasons:
            invalid_reasons.update(reasons)
        else:
            r.update(category=category or "<空白>", middle=middle or "<空白>", small=small or "<空白>", code=code, barcode=barcode, name=name, production=production, expiry=expiry, life=life, unit=unit, stock=stock, arrival=arrival, actual_days=(expiry-production).days, declared_days=life * {"D": 1, "M": 30, "Y": 365}[unit])
            usable.append(r)

    conflict_fields = ("商品条码", "商品名称", "保质期", "保质期单位", "是否该做临期折扣", "该批次累计到货数量")
    batch_conflicts = 0
    exact_duplicates = 0
    conflicting_keys = set()
    for key, group in batch_rows.items():
        signatures = {tuple(clean(row[f]) for f in conflict_fields) for row in group}
        if len(group) > 1:
            if len(signatures) == 1: exact_duplicates += len(group) - 1
            else:
                batch_conflicts += 1
                conflicting_keys.add(key)

    category_profile = {}
    for category, items in sorted(category_rows.items()):
        codes = {clean(x["商品编码"]) for x in items if clean(x["商品编码"])}
        keys = {(clean(x["商品编码"]), date_value(x["生产日期"]), date_value(x["有效日期"])) for x in items if clean(x["商品编码"]) and date_value(x["有效日期"])}
        expired = sum(1 for x in items if date_value(x["有效日期"]) and date_value(x["有效日期"]) < BUSINESS_DATE)
        category_profile[category] = {"rows": len(items), "products_by_product_code": len(codes), "batch_keys": len(keys), "expired_rows": expired}

    schemes = {name: collections.Counter() for name in SCHEMES}
    simulation_by_category = {cat: {"products_by_product_code": 0, "eligible_batches": 0, "expired_batches": 0, "executable_batches": 0, "stock_zero_batches": 0, "unable_to_classify_batches": 0, "schemes": {name: {"follow_up": 0, "history_baseline": 0} for name in SCHEMES}} for cat in category_profile}
    unique_usable = {}
    for r in usable:
        key = (r["code"], r["production"], r["expiry"])
        if key not in conflicting_keys and key not in unique_usable:
            unique_usable[key] = r
    for r in unique_usable.values():
        p = simulation_by_category[r["category"]]
        p["eligible_batches"] += 1
        if r["expiry"] < BUSINESS_DATE: p["expired_batches"] += 1
        if r["stock"] == 0: p["stock_zero_batches"] += 1
        if r["expiry"] < BUSINESS_DATE and r["stock"] != 0:
            for name, (ratio, low, high) in SCHEMES.items():
                days = window(r["actual_days"], ratio, low, high)
                if (BUSINESS_DATE - r["expiry"]).days <= days:
                    schemes[name]["follow_up"] += 1; p["schemes"][name]["follow_up"] += 1
                else:
                    schemes[name]["history_baseline"] += 1; p["schemes"][name]["history_baseline"] += 1

    history_by_window = {"pct_5_clamp_3_60": [], "pct_3_clamp_3_30": []}
    current_executable = []
    policy_coverage = {cat: {"calculable_batches": 0, "uncalculable_batches": 0, "calculable_product_codes": set(), "uncalculable_product_codes": set(), "reasons": collections.Counter()} for cat in category_profile}
    for r in unique_usable.values():
        policy_code, current = approved_policy(r)
        coverage = policy_coverage[r["category"]]
        if policy_code:
            coverage["calculable_batches"] += 1
            coverage["calculable_product_codes"].add(r["code"])
        else:
            coverage["uncalculable_batches"] += 1
            coverage["uncalculable_product_codes"].add(r["code"])
            coverage["reasons"][current] += 1
        if r["stock"] == 0:
            continue
        if policy_code and r["expiry"] < BUSINESS_DATE:
            overdue_days = (BUSINESS_DATE - r["expiry"]).days
            if overdue_days <= window(r["actual_days"], 0.05, 3, 60):
                history_by_window["pct_5_clamp_3_60"].append((r, "expired"))
            if overdue_days <= window(r["actual_days"], 0.03, 3, 30):
                history_by_window["pct_3_clamp_3_30"].append((r, "expired"))
        if policy_code and current in {"discount_50", "discount_20", "withdraw", "expired"} and r["expiry"] >= BUSINESS_DATE:
            current_executable.append((r, current))

    def workload(items):
        by_product = collections.defaultdict(list)
        for r, stage_name in items:
            by_product[r["code"]].append((r, stage_name))
        stage_counts = collections.Counter()
        batch_counts = []
        by_category = {}
        for code, product_items in by_product.items():
            highest = max((stage_name for _, stage_name in product_items), key=lambda x: {"discount_50": 1, "discount_20": 2, "withdraw": 3, "expired": 4}[x])
            stage_counts[highest] += 1
            batch_counts.append(len(product_items))
            category = product_items[0][0]["category"]
            if any(item[0]["category"] != category for item in product_items):
                raise SystemExit(f"Product {code} appears in multiple categories; workload scope is ambiguous.")
            bucket = by_category.setdefault(category, {"products": 0, "batches": 0, "stage_distribution": collections.Counter(), "batch_counts": []})
            bucket["products"] += 1
            bucket["batches"] += len(product_items)
            bucket["stage_distribution"][highest] += 1
            bucket["batch_counts"].append(len(product_items))
        def metrics(values):
            return {"average": sum(values) / len(values) if values else 0, "p50_nearest_rank": percentile_nearest_rank(values, .5), "p90_nearest_rank": percentile_nearest_rank(values, .9), "max": max(values) if values else 0}
        category_metrics = {}
        for cat in sorted(category_profile):
            value = by_category.get(cat, {"products": 0, "batches": 0, "stage_distribution": collections.Counter(), "batch_counts": []})
            category_metrics[cat] = {"products_and_open_tasks": value["products"], "batches": value["batches"], "stage_distribution": {name: value["stage_distribution"].get(name, 0) for name in ("discount_50", "discount_20", "withdraw", "expired")}, "batches_per_task": metrics(value["batch_counts"])}
        return {
            "batches": len(items), "products_and_open_tasks": len(by_product),
            "stage_distribution": {name: stage_counts.get(name, 0) for name in ("discount_50", "discount_20", "withdraw", "expired")},
            "batches_per_task": metrics(batch_counts),
            "by_category": category_metrics,
            "product_codes": sorted(by_product),
        }

    approved_history = history_by_window["pct_5_clamp_3_60"]
    history_workload = workload(approved_history)
    executable_workload = workload(current_executable)
    merged_by_key = {(r["code"], r["production"], r["expiry"]): (r, stage_name) for r, stage_name in approved_history}
    merged_by_key.update({(r["code"], r["production"], r["expiry"]): (r, stage_name) for r, stage_name in current_executable})
    merged_workload = workload(list(merged_by_key.values()))
    workload_result = {
        "approved_history_follow_up": history_workload,
        "current_executable": executable_workload,
        "merged_first_day": merged_workload,
        "overlapping_product_codes": len(set(history_workload.pop("product_codes")) & set(executable_workload.pop("product_codes"))),
    }
    merged_workload.pop("product_codes")

    scenario_stage_sets = {
        "A_all_current_stages": {"discount_50", "discount_20", "withdraw", "expired"},
        "B_ignore_existing_discount_50": {"discount_20", "withdraw", "expired"},
        "C_keep_withdraw_and_expired": {"withdraw", "expired"},
        "D_expired_only": {"expired"},
    }
    cold_start_scenarios = {}
    for history_name, history_items in history_by_window.items():
        history_summary = workload(history_items)
        history_summary.pop("product_codes")
        combinations = {}
        for scenario_name, included_stages in scenario_stage_sets.items():
            current_items = [(r, stage_name) for r, stage_name in current_executable if stage_name in included_stages]
            merged = {(r["code"], r["production"], r["expiry"]): (r, stage_name) for r, stage_name in history_items}
            merged.update({(r["code"], r["production"], r["expiry"]): (r, stage_name) for r, stage_name in current_items})
            summary = workload(list(merged.values()))
            summary.pop("product_codes")
            combinations[scenario_name] = {
                "first_day_open_product_tasks": summary["products_and_open_tasks"],
                "reduction_from_1474": 1474 - summary["products_and_open_tasks"],
                "stage_distribution": summary["stage_distribution"],
                "tasks_by_category": {category: value["products_and_open_tasks"] for category, value in summary["by_category"].items()},
            }
        cold_start_scenarios[history_name] = {
            "history_follow_up_batches": history_summary["batches"],
            "history_follow_up_product_tasks": history_summary["products_and_open_tasks"],
            "combinations": combinations,
        }
    if cold_start_scenarios["pct_5_clamp_3_60"]["combinations"]["A_all_current_stages"]["first_day_open_product_tasks"] != 1474:
        raise SystemExit("Approved A + 5% baseline no longer equals 1474.")

    def category_composition(category):
        items = [r for r in unique_usable.values() if r["category"] == category]
        buckets = collections.defaultdict(lambda: {"products": set(), "batches": 0, "shelf_life": collections.Counter()})
        for r in items:
            bucket = buckets[(r["middle"], r["small"])]
            bucket["products"].add(r["code"]); bucket["batches"] += 1
            bucket["shelf_life"][(r["life"], r["unit"])] += 1
        return [{"middle_category": middle, "small_category": small, "products": len(value["products"]), "batches": value["batches"], "shelf_life_distribution": {f"{life} {unit}": count for (life, unit), count in sorted(value["shelf_life"].items())}} for (middle, small), value in sorted(buckets.items())]

    gift_rows = []
    for r in sorted((x for x in unique_usable.values() if x["category"] == "赠品小样"), key=lambda x: x["name"]):
        gift_rows.append({"middle_category": r["middle"], "small_category": r["small"], "product_name": r["name"], "shelf_life": r["life"], "unit": r["unit"], "declared_days": r["declared_days"]})

    long_life_boundary = {}
    for category in sorted(GENERAL_LONG_LIFE_CATEGORIES):
        buckets = {"gt_6_months": {"products": set(), "batches": 0}, "eq_6_months": {"products": set(), "batches": 0}, "lt_6_months": {"products": set(), "batches": 0}}
        short_detail = collections.defaultdict(lambda: {"products": set(), "batches": 0})
        items = [r for r in unique_usable.values() if r["category"] == category]
        for r in items:
            key = "gt_6_months" if r["declared_days"] > 180 else "eq_6_months" if r["declared_days"] == 180 else "lt_6_months"
            buckets[key]["products"].add(r["code"]); buckets[key]["batches"] += 1
            if key != "gt_6_months":
                detail = short_detail[(r["middle"], r["small"])]
                detail["products"].add(r["code"]); detail["batches"] += 1
        long_life_boundary[category] = {
            "total_products": len({r["code"] for r in items}), "total_batches": len(items),
            **{name: {"products": len(value["products"]), "batches": value["batches"]} for name, value in buckets.items()},
            "le_6_months_distribution": [{"middle_category": middle, "small_category": small, "products": len(value["products"]), "batches": value["batches"]} for (middle, small), value in sorted(short_detail.items())],
        }

    coverage_result = {category: {"calculable_batches": value["calculable_batches"], "uncalculable_batches": value["uncalculable_batches"], "calculable_products": len(value["calculable_product_codes"]), "uncalculable_products": len(value["uncalculable_product_codes"]), "uncalculable_reasons": counter_dict(value["reasons"])} for category, value in sorted(policy_coverage.items())}
    usable_row_numbers = {r["row"] for r in usable}
    for cat, p in simulation_by_category.items():
        p["products_by_product_code"] = category_profile[cat]["products_by_product_code"]
        invalid_rows = sum(1 for r in rows if (clean(r["商品大类"]) or "<空白>") == cat and r["row"] not in usable_row_numbers)
        conflict_keys_in_category = {key for key in conflicting_keys if (clean(batch_rows[key][0]["商品大类"]) or "<空白>") == cat}
        p["unable_to_classify_batches"] = invalid_rows + len(conflict_keys_in_category)
    life_days = [r["actual_days"] for r in usable]
    life_by_unit = collections.Counter(r["unit"] for r in usable)
    result = {
        "source": {"path": str(source), "bytes": size, "sha256": before, "business_date": BUSINESS_DATE.isoformat()},
        "workbook": {"sheets": sheets, "required_headers_missing": missing},
        "totals": {"source_rows": len(rows), "unique_product_codes": len(product_rows), "candidate_batch_keys": len(batch_rows), "usable_unique_batches": len(unique_usable), "usable_rows": len(usable), "unusable_rows": len(rows)-len(usable), "exact_duplicate_rows": exact_duplicates, "conflicting_batch_keys": batch_conflicts, "nonunique_barcodes": sum(1 for v in barcode_rows.values() if len({clean(x["商品编码"]) for x in v}) > 1)},
        "field_quality": {k: counter_dict(v) for k, v in field_quality.items()},
        "invalid_reasons": counter_dict(invalid_reasons),
        "categories": category_profile,
        "identity_quality": {"product_codes_with_multiple_names": sum(1 for v in product_rows.values() if len({clean(x["商品名称"]) for x in v}) > 1), "product_codes_with_multiple_barcodes": sum(1 for v in product_rows.values() if len({clean(x["商品条码"]) for x in v}) > 1)},
        "shelf_life_actual_days": {"count": len(life_days), "min": min(life_days) if life_days else None, "max": max(life_days) if life_days else None, "median": sorted(life_days)[len(life_days)//2] if life_days else None, "unit_counts": counter_dict(life_by_unit)},
        "simulation": {"eligible_unique_batch_basis": "product_code + production_date + expiry_date; first row only for identical keys", "schemes": {name: dict(value) for name, value in schemes.items()}, "by_category": simulation_by_category, "stage_window_reverse": "existing authoritative calculation: D/M/Y to 1/30/365 days; total >270 then 90/60/14 else 30/14/7; not a historical-expiry follow-up rule"},
        "company_stage_policy": {
            "authority": "Product-manager supplied company rule; neutral Excel discount date is not used for policy inference.",
            "month_conversion_for_simulation_only": "1 month = 30 days; total shelf-life D/M/Y = 1/30/365. This is an analysis convention aligned with existing code, not an additional company rule.",
            "policy_groups": {
                "food_expiry_v1": "食品: total <=270 days 30/14/7; >270 days 90/60/14",
                "pet_expiry_v1": "宠物: 90/60/14 regardless of total shelf-life",
                "general_long_expiry_v1": "日用/美妆/家居/香氛香水/文具/潮流玩具: total >180 days only, 180/90/14",
                "excluded_v1": "应季搭配/赠品小样 import normally but do not participate in expiry management in V1",
                "unresolved": "six general categories with total <=180 days have no supplied policy and create no expiry tasks",
            },
            "coverage_by_category": coverage_result,
            "seasonal_matching": {"category": "应季搭配", "composition": category_composition("应季搭配"), "decision": "unresolved: source has no reliable season or off-season approval attribute"},
            "gift_sample_matching": {"category": "赠品小样", "products": gift_rows, "decision": "unresolved: source category and attributes do not prove an original company policy"},
            "general_long_life_boundary": long_life_boundary,
        },
        "first_day_workload_company_policy": workload_result,
        "cold_start_scenarios": cold_start_scenarios,
    }
    after = hashlib.sha256(source.read_bytes()).hexdigest().upper()
    if after != before or source.stat().st_size != size:
        raise SystemExit("Source changed during read; no output accepted.")
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"source_sha256": before, "source_bytes": size, "usable_unique_batches": len(unique_usable), "output": str(output)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
